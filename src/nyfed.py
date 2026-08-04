"""
NY Fed Household Debt and Credit collector.

Data source: Federal Reserve Bank of New York, Consumer Credit Panel (based on Equifax data).
Report: Household Debt and Credit — quarterly Excel workbook published at
https://www.newyorkfed.org/microeconomics/hhdc

Two sheets are extracted:

- "Page 3 Data" — total debt balances by category in trillions of dollars going
  back to Q1 1999. Categories: Mortgage, HE Revolving (HELOC), Auto Loan,
  Credit Card, Student Loan, Other.
- "Page 12 Data" — percent of balance 90+ days delinquent by loan type going
  back to Q1 2003. Columns: MORTGAGE, HELOC, AUTO, CC, STUDENT LOAN, OTHER, ALL.

"Other" is a catch-all that includes medical debt, personal loans, and retail
financing — it is the only publicly available proxy for medical debt at this
time granularity.

Balance CSVs use millions of dollars (consistent with FRED series) so the
viz layer can treat all debt series uniformly. Delinquency CSVs are in percent.
"""

from datetime import date
from pathlib import Path
import json
import re

import polars as pl
import requests


BASE_URL = "https://www.newyorkfed.org/medialibrary/interactives/householdcredit/data/xls"
FILENAME_PATTERN = "HHD_C_Report_{year}Q{quarter}.xlsx"

# Output CSV names → column name in the Excel sheet (exact match preferred,
# case-insensitive substring as fallback)
SERIES = {
    "nyfed_mortgage":    "Mortgage",
    "nyfed_he_revolving": "HE Revolving",
    "nyfed_auto":        "Auto Loan",
    "nyfed_credit_card": "Credit Card",
    "nyfed_student":     "Student Loan",
    "nyfed_other":       "Other",
    "nyfed_total":       "Total",
}

DELINQ_SERIES = {
    "nyfed_delinq_mortgage":     "MORTGAGE",
    "nyfed_delinq_he_revolving": "HELOC",
    "nyfed_delinq_auto":         "AUTO",
    "nyfed_delinq_credit_card":  "CC",
    "nyfed_delinq_student":      "STUDENT LOAN",
    "nyfed_delinq_other":        "OTHER",
    "nyfed_delinq_total":        "ALL",
}

# Sheet configs: how to locate each sheet and scale its values
BALANCE_SHEET = {
    "name_pattern": r"page.?3\b",
    "title_hint": "Total Debt Balance",
    "series_map": SERIES,
    # Values are in trillions → convert to millions
    "scale": 1_000_000,
}
DELINQ_SHEET = {
    "name_pattern": r"page.?12\b",
    "title_hint": "90+ Days Delinquent by Loan Type",
    "series_map": DELINQ_SERIES,
    # Values are already in percent
    "scale": 1,
}

_NYFED_COMMON = {
    "frequency": "Quarterly",
    "seasonal_adjustment": "Not Seasonally Adjusted",
    "source": "NY Fed / Equifax Consumer Credit Panel",
    "source_url": "https://www.newyorkfed.org/microeconomics/hhdc",
}

METADATA = {
    "nyfed_mortgage":     {"title": "NY Fed: Mortgage Debt Balance",             "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_he_revolving": {"title": "NY Fed: Home Equity Revolving (HELOC) Balance", "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_auto":         {"title": "NY Fed: Auto Loan Balance",                 "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_credit_card":  {"title": "NY Fed: Credit Card Balance",               "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_student":      {"title": "NY Fed: Student Loan Balance",              "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_other":        {"title": "NY Fed: Other Debt Balance (incl. medical)", "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_total":        {"title": "NY Fed: Total Household Debt Balance",       "units": "Millions of U.S. Dollars", **_NYFED_COMMON},
    "nyfed_delinq_mortgage":     {"title": "NY Fed: Mortgage Balance 90+ Days Delinquent",      "units": "Percent of Balance", **_NYFED_COMMON},
    "nyfed_delinq_he_revolving": {"title": "NY Fed: HELOC Balance 90+ Days Delinquent",         "units": "Percent of Balance", **_NYFED_COMMON},
    "nyfed_delinq_auto":         {"title": "NY Fed: Auto Loan Balance 90+ Days Delinquent",     "units": "Percent of Balance", **_NYFED_COMMON},
    "nyfed_delinq_credit_card":  {"title": "NY Fed: Credit Card Balance 90+ Days Delinquent",   "units": "Percent of Balance", **_NYFED_COMMON},
    "nyfed_delinq_student":      {"title": "NY Fed: Student Loan Balance 90+ Days Delinquent",  "units": "Percent of Balance", **_NYFED_COMMON},
    "nyfed_delinq_other":        {"title": "NY Fed: Other Debt Balance 90+ Days Delinquent",    "units": "Percent of Balance", **_NYFED_COMMON},
    "nyfed_delinq_total":        {"title": "NY Fed: All Debt Balance 90+ Days Delinquent",      "units": "Percent of Balance", **_NYFED_COMMON},
}


class NyFedCollector:
    def __init__(self, output_dir: str = "data/raw"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.metadata_file = self.output_dir.parent / "metadata.json"

    def _latest_report_url(self) -> tuple[str, str]:
        """Return (url, label) for the most recent available quarterly report."""
        today = date.today()
        # Current quarter — NY Fed publishes ~5–6 weeks after quarter-end,
        # so subtract one quarter as the most likely published quarter.
        current_q = (today.month - 1) // 3 + 1
        current_year = today.year

        # Build candidates starting one quarter back from today
        candidates = []
        y, q = current_year, current_q - 1
        if q == 0:
            y -= 1
            q = 4
        for _ in range(8):  # check up to 8 quarters back
            candidates.append((y, q))
            q -= 1
            if q == 0:
                y -= 1
                q = 4

        for year, quarter in candidates:
            filename = FILENAME_PATTERN.format(year=year, quarter=quarter)
            url = f"{BASE_URL}/{filename}"
            try:
                r = requests.head(url, timeout=10, allow_redirects=True,
                                  headers={"User-Agent": "Mozilla/5.0"})
                ct = r.headers.get("Content-Type", "")
                # xlsx files are application/vnd.openxmlformats... or application/zip/octet-stream
                if r.status_code == 200 and ("spreadsheet" in ct or "zip" in ct or "octet" in ct or "excel" in ct):
                    return url, f"{year}Q{quarter}"
            except requests.RequestException:
                continue

        # Fall back: try downloading each candidate and check magic bytes
        for year, quarter in candidates:
            filename = FILENAME_PATTERN.format(year=year, quarter=quarter)
            url = f"{BASE_URL}/{filename}"
            try:
                r = requests.get(url, timeout=30, stream=True,
                                 headers={"User-Agent": "Mozilla/5.0"})
                first_bytes = next(r.iter_content(4), b"")
                r.close()
                # xlsx (zip) magic: PK\x03\x04
                if first_bytes[:2] == b"PK":
                    return url, f"{year}Q{quarter}"
            except requests.RequestException:
                continue

        raise RuntimeError("Could not find a valid NY Fed report. Try: python main.py --source nyfed --nyfed-quarter 2025Q1")

    def _download_workbook(self, url: str):
        """Download the Excel workbook and return as bytes."""
        print(f"⬇️  Downloading {url} ...")
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        return r.content

    def _find_sheet(self, wb, name_pattern: str, title_hint: str):
        """Locate a data sheet by name pattern, falling back to a title search."""
        for name in wb.sheetnames:
            if re.search(name_pattern, name, re.IGNORECASE):
                return wb[name]
        # Fall back: first worksheet whose top rows contain the title hint
        for name in wb.sheetnames:
            ws = wb[name]
            if not hasattr(ws, "iter_rows"):  # skip chartsheets
                continue
            header_text = " ".join(
                str(ws.cell(r, c).value or "")
                for r in range(1, 6)
                for c in range(1, 15)
            )
            if title_hint.lower() in header_text.lower():
                return ws
        return None

    def _parse_sheet(self, wb, config: dict) -> pl.DataFrame:
        """Parse a quarterly data sheet into a tidy DataFrame per its config."""
        series_map = config["series_map"]
        sheet = self._find_sheet(wb, config["name_pattern"], config["title_hint"])
        if sheet is None:
            available = ", ".join(wb.sheetnames)
            raise RuntimeError(
                f"Could not locate sheet matching '{config['title_hint']}'. Available sheets: {available}"
            )

        # Read all rows into a list-of-lists
        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]

        # Locate header row (contains at least half the expected column labels)
        header_idx = None
        for i, row in enumerate(rows):
            cells = [str(v or "").strip() for v in row]
            hits = sum(
                any(label.lower() == c.lower() or label.lower() in c.lower() for c in cells)
                for label in series_map.values()
            )
            if hits >= len(series_map) / 2:
                header_idx = i
                break
        if header_idx is None:
            raise RuntimeError(f"Could not find header row for '{config['title_hint']}'")

        headers = [str(v or "").strip() for v in rows[header_idx]]

        # Map column names → indices; exact match wins over substring match
        col_map = {}
        for series_key, col_label in series_map.items():
            for idx, h in enumerate(headers):
                if col_label.lower() == h.lower():
                    col_map[series_key] = idx
                    break
            else:
                for idx, h in enumerate(headers):
                    if col_label.lower() in h.lower():
                        col_map[series_key] = idx
                        break

        # Find date column (first column that looks like a quarter label e.g. "99:Q1")
        date_col = 0  # usually column 0

        records = []
        for row in rows[header_idx + 1:]:
            if not row or row[date_col] is None:
                continue
            raw_date = str(row[date_col]).strip()
            # Parse formats: "99:Q1", "2003:Q1", "Q1 1999", "1999Q1"
            parsed = _parse_quarter_date(raw_date)
            if parsed is None:
                continue
            record = {"date": parsed}
            for series_key, col_idx in col_map.items():
                try:
                    val = row[col_idx]
                    if val is not None:
                        record[series_key] = float(val) * config["scale"]
                except (TypeError, ValueError):
                    pass
            records.append(record)

        return pl.DataFrame(records)

    def save_metadata(self, series_id: str, extra: dict = None):
        if self.metadata_file.exists():
            with open(self.metadata_file) as f:
                all_meta = json.load(f)
        else:
            all_meta = {}
        entry = dict(METADATA[series_id])
        entry["last_updated"] = date.today().isoformat()
        if extra:
            entry.update(extra)
        all_meta[series_id] = entry
        with open(self.metadata_file, "w") as f:
            json.dump(all_meta, f, indent=2)

    def _write_series(self, df: pl.DataFrame, series_map: dict):
        for series_key in series_map:
            if series_key not in df.columns:
                print(f"⚠️  Column missing for {series_key}, skipping")
                continue
            out = df.select(["date", series_key]).rename({series_key: "value"}).drop_nulls()
            filepath = self.output_dir / f"{series_key}.csv"
            out.write_csv(filepath)
            self.save_metadata(series_key)
            print(f"✅ Saved {series_key}.csv ({len(out)} rows, {out['date'].min()} to {out['date'].max()})")

    def collect_all(self, quarter: str = None):
        if quarter:
            m = re.match(r"(\d{4})Q(\d)", quarter.upper())
            if not m:
                raise ValueError(f"Invalid quarter format '{quarter}'. Use e.g. 2024Q4")
            year, q = int(m.group(1)), int(m.group(2))
            filename = FILENAME_PATTERN.format(year=year, quarter=q)
            url = f"{BASE_URL}/{filename}"
            label = quarter.upper()
        else:
            url, label = self._latest_report_url()

        print(f"📊 NY Fed Household Debt Report — {label}")
        content = self._download_workbook(url)

        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)

        for config in (BALANCE_SHEET, DELINQ_SHEET):
            try:
                df = self._parse_sheet(wb, config)
            except RuntimeError as e:
                print(f"❌ {e}")
                continue
            if df.is_empty():
                print(f"❌ Parsed DataFrame for '{config['title_hint']}' is empty — check sheet structure")
                continue
            self._write_series(df, config["series_map"])


def _parse_quarter_date(raw: str) -> str | None:
    """Convert quarter labels like '99:Q1', '2003:Q1', 'Q1 2003' to ISO date strings."""
    raw = raw.strip()
    # "YY:Q#" or "YYYY:Q#"
    m = re.match(r"(\d{2,4}):Q(\d)", raw)
    if m:
        year, quarter = int(m.group(1)), int(m.group(2))
        if year < 100:
            year += 1900 if year >= 99 else 2000
        return _quarter_to_date(year, quarter)
    # "Q# YYYY" or "Q#YYYY"
    m = re.match(r"Q(\d)\s*(\d{4})", raw, re.IGNORECASE)
    if m:
        return _quarter_to_date(int(m.group(2)), int(m.group(1)))
    # "YYYYQ#"
    m = re.match(r"(\d{4})Q(\d)", raw, re.IGNORECASE)
    if m:
        return _quarter_to_date(int(m.group(1)), int(m.group(2)))
    return None


def _quarter_to_date(year: int, quarter: int) -> str:
    """Return the last month of the quarter as ISO YYYY-MM-01."""
    month = quarter * 3
    return f"{year}-{month:02d}-01"
