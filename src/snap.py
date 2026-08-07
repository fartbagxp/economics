"""
SNAP (Supplemental Nutrition Assistance Program) participation collector.

Data source: USDA Food and Nutrition Service, National Data Bank.
Historical archive: https://www.fna.usda.gov/pd/supplemental-nutrition-assistance-program-snap

Downloads the "FY69 to current" zip of per-fiscal-year Excel workbooks and
extracts the "US Summary" sheet from each, which reports national monthly
participation (households and persons). Fiscal years 1969-1988 are national-only
in a different, hand-formatted layout and are skipped — FY1989 onward already
gives 35+ years of consistent monthly history.
"""

from datetime import date
from pathlib import Path
import io
import json
import re
import zipfile

import polars as pl
import requests


ZIP_URL = "https://www.fna.usda.gov/sites/default/files/resource-files/snap-zip-fy69tocurrent-7.zip"

MONTH_RE = re.compile(
    r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+(\d{2,4})$",
    re.IGNORECASE,
)
MONTH_NUM = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

METADATA = {
    "title": "SNAP: Persons Participating (National)",
    "units": "Persons",
    "frequency": "Monthly",
    "seasonal_adjustment": "Not Seasonally Adjusted",
    "source": "USDA Food and Nutrition Service, National Data Bank",
    "source_url": "https://www.fna.usda.gov/pd/supplemental-nutrition-assistance-program-snap",
}


def _parse_month_label(raw: str) -> str | None:
    """Convert labels like 'Oct 2009' or 'July 68' to an ISO date (first of month)."""
    if not raw:
        return None
    m = MONTH_RE.match(raw.strip())
    if not m:
        return None
    month = MONTH_NUM[m.group(1)[:3].lower()]
    year = int(m.group(2))
    if year < 100:
        year += 1900 if year >= 70 else 2000
    return f"{year}-{month:02d}-01"


def _find_us_summary_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "us summary":
            return wb[name]
    return None


class SnapCollector:
    """Collector for USDA SNAP national participation data."""

    def __init__(self, output_dir: str = "data/raw"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.metadata_file = self.output_dir.parent / "metadata.json"

    def _download_zip(self) -> bytes:
        print(f"⬇️  Downloading {ZIP_URL} ...")
        r = requests.get(ZIP_URL, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        return r.content

    def _parse_workbook(self, name: str, content: bytes) -> list[dict]:
        """Parse one fiscal-year workbook's US Summary sheet into month/persons records."""
        records = []
        if name.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            sheet = _find_us_summary_sheet(wb)
            if sheet is None:
                return records
            for row in sheet.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                iso_date = _parse_month_label(str(row[0]))
                if iso_date is None or len(row) < 3 or row[2] is None:
                    continue
                try:
                    records.append({"date": iso_date, "value": float(row[2])})
                except (TypeError, ValueError):
                    continue
        elif name.lower().endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            sheet_name = next(
                (s for s in wb.sheet_names() if s.strip().lower() == "us summary"), None
            )
            if sheet_name is None:
                return records
            sheet = wb.sheet_by_name(sheet_name)
            for r in range(sheet.nrows):
                row = sheet.row_values(r)
                if not row or not row[0]:
                    continue
                iso_date = _parse_month_label(str(row[0]))
                if iso_date is None or len(row) < 3 or row[2] in (None, ""):
                    continue
                try:
                    records.append({"date": iso_date, "value": float(row[2])})
                except (TypeError, ValueError):
                    continue
        return records

    def save_metadata(self):
        if self.metadata_file.exists():
            with open(self.metadata_file) as f:
                all_meta = json.load(f)
        else:
            all_meta = {}
        entry = dict(METADATA)
        entry["last_updated"] = date.today().isoformat()
        all_meta["snap_persons"] = entry
        with open(self.metadata_file, "w") as f:
            json.dump(all_meta, f, indent=2)

    def collect_all(self):
        content = self._download_zip()
        zf = zipfile.ZipFile(io.BytesIO(content))

        all_records = []
        for name in zf.namelist():
            base = name.rsplit("/", 1)[-1]
            if not re.match(r"^FY\d{2}\.xlsx?$", base, re.IGNORECASE):
                continue  # skip the special-format 1969-88 national-only file
            records = self._parse_workbook(base, zf.read(name))
            all_records.extend(records)

        if not all_records:
            raise RuntimeError("No SNAP records parsed from any fiscal-year workbook")

        df = (
            pl.DataFrame(all_records)
            .with_columns(pl.col("date").str.to_date())
            .unique(subset="date")
            .sort("date")
        )

        filepath = self.output_dir / "snap_persons.csv"
        df.write_csv(filepath)
        self.save_metadata()
        print(f"✅ Saved snap_persons.csv ({len(df)} rows, {df['date'].min()} to {df['date'].max()})")
        return df
